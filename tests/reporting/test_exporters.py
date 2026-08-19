"""Ihracatci (CSV / Excel / PDF) ve rapor sorgusu testleri.

Dosya ciktilari ``tmp_path`` altina yazilir: ihracatcilar dosya yolunu
:data:`app.core.paths.EXPORT_DIR` altinda cozdugu icin, testler once bu
koku gecici klasore tasir. Boylece hicbir test gercek ``exports/``
klasorune dokunmaz.

Bos veri senaryosu bilincli olarak her bicimde ayri ayri dogrulanir:
raporlarin coktugu an, genellikle hicbir kaydin olmadigi andir.
"""

from __future__ import annotations

import codecs
import csv
import re
from datetime import UTC, date, datetime
from decimal import Decimal

import pytest
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core import paths
from app.core.exceptions import NotFoundError, ValidationError
from app.domain.enums import (
    ChargeType,
    Currency,
    HousekeepingTaskType,
    MaintenanceCategory,
    PaymentMethod,
    ReservationSource,
    ReservationStatus,
    RoomHousekeepingStatus,
    TransactionDirection,
)
from app.domain.value_objects import DateRange, Money
from app.infrastructure.db.models import (
    CashRegisterEntry,
    Charge,
    Folio,
    HousekeepingTask,
    InventoryItem,
    MaintenanceTicket,
    Payment,
    Reservation,
    ReservationRoom,
)
from app.reporting import queries
from app.reporting.exporters import (
    export_csv,
    export_excel,
    export_pdf,
    get_exporter,
)
from app.reporting.exporters.csv_exporter import CSV_DELIMITER
from app.reporting.exporters.pdf_exporter import (
    ReportFonts,
    register_report_fonts,
    sanitize_text,
)
from app.reporting.models import (
    EMPTY_TABLE_MESSAGE,
    ReportColumn,
    ReportTable,
)

#: Test donemi: 10-13 Agustos 2026 (3 gece).
DONEM = DateRange(date(2026, 8, 10), date(2026, 8, 13))

#: Yapay zeka seffaflik dipnotu - ucu de bunu tasiyabilmelidir.
AI_NOTU = "Bu rapor yapay zeka tarafindan olusturulmustur."


# ==========================================================================
#  Fiksturler
# ==========================================================================
@pytest.fixture
def export_dir(tmp_path, monkeypatch):
    """Disa aktarma kokunu gecici klasore tasir."""
    data_root = tmp_path / "veri"
    exports = data_root / "exports"
    exports.mkdir(parents=True)
    monkeypatch.setattr(paths, "DATA_ROOT", data_root)
    monkeypatch.setattr(paths, "EXPORT_DIR", exports)
    return exports


@pytest.fixture
def turkce_tablo() -> ReportTable:
    """Turkce karakter, para ve tarih iceren ornek tablo (tamamen uydurma)."""
    columns = [
        ReportColumn("ucret_turu", "Ucret Turu", align="left"),
        ReportColumn("aciklama", "Aciklama", align="left"),
        ReportColumn("gun", "Tarih", align="center", format="date"),
        ReportColumn("adet", "Adet", align="right", format="integer"),
        ReportColumn("tutar", "Tutar", align="right", format="money"),
    ]
    rows = [
        {
            "ucret_turu": "Konaklama Ücreti",
            "aciklama": "Şahin Yıldız Apart - oda geliri",
            "gun": date(2026, 8, 10),
            "adet": 3,
            "tutar": Money.of("1234.56"),
        },
        {
            "ucret_turu": "Çamaşırhane",
            "aciklama": "Ütü & yıkama <hizmeti>",
            "gun": date(2026, 8, 11),
            "adet": 12,
            "tutar": Money.of("890.10"),
        },
        {
            "ucret_turu": "Otopark",
            "aciklama": "Günlük ücret",
            "gun": date(2026, 8, 12),
            "adet": 1,
            "tutar": Money.of("150.00"),
        },
    ]
    return ReportTable(
        title="Ücret Türü Bazında Gelir",
        columns=columns,
        rows=rows,
        subtitle="Deneme Oteli",
        filters_description=DONEM.format(),
        footer_note=AI_NOTU,
    )


@pytest.fixture
def bos_tablo() -> ReportTable:
    """Hicbir satiri olmayan tablo - bos veri senaryosu."""
    return ReportTable(
        title="Bos Rapor",
        columns=[
            ReportColumn("gun", "Tarih", format="date"),
            ReportColumn("tutar", "Tutar", align="right", format="money"),
        ],
        rows=[],
        subtitle="Deneme Oteli",
        footer_note=AI_NOTU,
    )


@pytest.fixture
def dolu_veri(session: Session, sample_property, sample_rooms, sample_guest):
    """Uc gecelik tek rezervasyon, iki folyo, ucretler ve bir tahsilat.

    Tum degerler uydurmadir; gercek bir kisiye veya isletmeye ait degildir.
    """
    reservation = Reservation(
        property_id=sample_property.id,
        confirmation_number="TEST-000001",
        status=ReservationStatus.CHECKED_IN,
        source=ReservationSource.BOOKING_COM,
        primary_guest_id=sample_guest.id,
        check_in_date=date(2026, 8, 10),
        check_out_date=date(2026, 8, 13),
        adults=2,
        currency=Currency.TRY,
    )
    session.add(reservation)
    session.flush()

    res_room = ReservationRoom(
        reservation_id=reservation.id,
        room_type_id=sample_rooms[0].room_type_id,
        room_id=sample_rooms[0].id,
        check_in_date=date(2026, 8, 10),
        check_out_date=date(2026, 8, 13),
        adults=2,
        nightly_rate=Decimal("1000.00"),
        total_amount=Decimal("3000.00"),
    )
    session.add(res_room)
    session.flush()

    folio = Folio(
        property_id=sample_property.id,
        folio_number="F-000001",
        reservation_id=reservation.id,
        reservation_room_id=res_room.id,
        guest_id=sample_guest.id,
    )
    # Ikinci folyo bilerek oda satirina baglanmaz: oda tipi kirilimindeki
    # "ilk oda satirini varsay" yolunu da sinamak icin.
    ek_folio = Folio(
        property_id=sample_property.id,
        folio_number="F-000002",
        reservation_id=reservation.id,
        guest_id=sample_guest.id,
    )
    session.add_all([folio, ek_folio])
    session.flush()

    for gun in (date(2026, 8, 10), date(2026, 8, 11), date(2026, 8, 12)):
        charge = Charge(
            folio_id=folio.id,
            charge_type=ChargeType.ROOM,
            description=f"Konaklama - {gun.strftime('%d.%m.%Y')}",
            charge_date=gun,
            quantity=Decimal("1.000"),
            unit_price=Decimal("1000.00"),
            tax_rate_percent=Decimal("0.00"),
        )
        charge.compute_totals()
        session.add(charge)

    restoran = Charge(
        folio_id=folio.id,
        charge_type=ChargeType.RESTAURANT,
        description="Akşam yemeği",
        charge_date=date(2026, 8, 11),
        quantity=Decimal("2.000"),
        unit_price=Decimal("250.00"),
        tax_rate_percent=Decimal("0.00"),
    )
    restoran.compute_totals()

    spa = Charge(
        folio_id=ek_folio.id,
        charge_type=ChargeType.SPA,
        description="Masaj",
        charge_date=date(2026, 8, 11),
        quantity=Decimal("1.000"),
        unit_price=Decimal("300.00"),
        tax_rate_percent=Decimal("0.00"),
    )
    spa.compute_totals()

    iptal_edilen = Charge(
        folio_id=folio.id,
        charge_type=ChargeType.ROOM,
        description="Yanlis islenen oda ucreti",
        charge_date=date(2026, 8, 11),
        quantity=Decimal("1.000"),
        unit_price=Decimal("9999.00"),
        tax_rate_percent=Decimal("0.00"),
    )
    iptal_edilen.compute_totals()
    iptal_edilen.void("Hatali kayit")

    session.add_all([restoran, spa, iptal_edilen])

    session.add(
        Payment(
            folio_id=folio.id,
            method=PaymentMethod.CASH,
            amount=Decimal("2000.00"),
            paid_at=datetime(2026, 8, 12, 9, 30, tzinfo=UTC),
        )
    )
    session.add(
        CashRegisterEntry(
            property_id=sample_property.id,
            entry_date=date(2026, 8, 12),
            direction=TransactionDirection.INCOME,
            category="Oda Geliri",
            description="Nakit tahsilat",
            amount=Decimal("2000.00"),
        )
    )
    session.add(
        HousekeepingTask(
            property_id=sample_property.id,
            room_id=sample_rooms[0].id,
            task_type=HousekeepingTaskType.DAILY_CLEANING,
            scheduled_date=date(2026, 8, 11),
            estimated_minutes=25,
        )
    )
    session.add(
        MaintenanceTicket(
            property_id=sample_property.id,
            ticket_number="ARZ-0001",
            room_id=sample_rooms[1].id,
            category=MaintenanceCategory.PLUMBING,
            title="Lavabo akıtıyor",
            description="Banyo lavabosunun sifonu damlatıyor.",
            reported_at=datetime(2026, 8, 11, 8, 0, tzinfo=UTC),
            labor_cost=Decimal("150.00"),
        )
    )
    session.add(
        InventoryItem(
            property_id=sample_property.id,
            sku="MNB-001",
            name="Maden suyu",
            category="Minibar",
            current_stock=Decimal("4.000"),
            minimum_stock=Decimal("10.000"),
            unit_cost=Decimal("12.50"),
        )
    )
    session.commit()
    return {"reservation": reservation, "folio": folio, "res_room": res_room}


# ==========================================================================
#  CSV
# ==========================================================================
def test_csv_bom_ile_yazilir(export_dir, turkce_tablo):
    """BOM olmadan Excel dosyayi cp1254 sanir ve Turkce harfler bozulur."""
    hedef = export_csv(turkce_tablo, "gelir.csv")
    ham = hedef.read_bytes()
    assert ham.startswith(codecs.BOM_UTF8)


def test_csv_turkce_karakterler_dogru_okunur(export_dir, turkce_tablo):
    hedef = export_csv(turkce_tablo, "gelir.csv")
    metin = hedef.read_text(encoding="utf-8-sig")
    assert "Çamaşırhane" in metin
    assert "Şahin Yıldız Apart" in metin
    assert "Günlük ücret" in metin


def test_csv_noktali_virgul_ayirici_kullanir(export_dir, turkce_tablo):
    hedef = export_csv(turkce_tablo, "gelir.csv")
    with hedef.open(encoding="utf-8-sig", newline="") as handle:
        satirlar = list(csv.reader(handle, delimiter=CSV_DELIMITER))
    assert satirlar[0] == turkce_tablo.header_titles
    assert len(satirlar[1]) == turkce_tablo.column_count


def test_csv_para_virgullu_yazilir_ve_bolunmez(export_dir, turkce_tablo):
    """1.234,56 tutari noktali virgul sayesinde tek hucrede kalir."""
    hedef = export_csv(turkce_tablo, "gelir.csv")
    with hedef.open(encoding="utf-8-sig", newline="") as handle:
        satirlar = list(csv.reader(handle, delimiter=CSV_DELIMITER))
    assert "1.234,56" in satirlar[1][-1]


def test_csv_bos_tabloda_gecerli_dosya_uretir(export_dir, bos_tablo):
    hedef = export_csv(bos_tablo, "bos.csv")
    with hedef.open(encoding="utf-8-sig", newline="") as handle:
        satirlar = list(csv.reader(handle, delimiter=CSV_DELIMITER))
    assert satirlar[0] == ["Tarih", "Tutar"]
    assert satirlar[1][0] == EMPTY_TABLE_MESSAGE
    assert hedef.stat().st_size > 0


def test_csv_dipnotu_yazar(export_dir, turkce_tablo):
    hedef = export_csv(turkce_tablo, "gelir.csv")
    metin = hedef.read_text(encoding="utf-8-sig")
    assert AI_NOTU in metin


# ==========================================================================
#  Excel
# ==========================================================================
def _baslik_satiri(sheet, ilk_baslik: str) -> int:
    """Sutun basliklarinin bulundugu satir numarasini bulur."""
    for row in sheet.iter_rows():
        if row[0].value == ilk_baslik:
            return row[0].row
    raise AssertionError(f"Baslik satiri bulunamadi: {ilk_baslik}")


def test_excel_acilabilir_ve_basliklar_dogru(export_dir, turkce_tablo):
    hedef = export_excel(turkce_tablo, "gelir.xlsx")
    sheet = load_workbook(hedef).active
    baslik = _baslik_satiri(sheet, "Ucret Turu")
    basliklar = [cell.value for cell in sheet[baslik]]
    assert basliklar == turkce_tablo.header_titles


def test_excel_para_sayisal_yazilir(export_dir, turkce_tablo):
    """Tutar metin degil sayi olmalidir; aksi halde Excel'de toplam alinamaz."""
    hedef = export_excel(turkce_tablo, "gelir.xlsx")
    sheet = load_workbook(hedef).active
    baslik = _baslik_satiri(sheet, "Ucret Turu")
    hucre = sheet.cell(row=baslik + 1, column=5)
    assert isinstance(hucre.value, (int, float, Decimal))
    assert float(hucre.value) == pytest.approx(1234.56)
    assert "#,##0.00" in hucre.number_format


def test_excel_dondurulmus_satir_ve_oto_filtre(export_dir, turkce_tablo):
    hedef = export_excel(turkce_tablo, "gelir.xlsx")
    sheet = load_workbook(hedef).active
    baslik = _baslik_satiri(sheet, "Ucret Turu")
    assert sheet.freeze_panes == f"A{baslik + 1}"
    assert sheet.auto_filter.ref is not None
    assert sheet.auto_filter.ref.startswith(f"A{baslik}")


def test_excel_sutun_genislikleri_ayarlanir(export_dir, turkce_tablo):
    hedef = export_excel(turkce_tablo, "gelir.xlsx")
    sheet = load_workbook(hedef).active
    assert sheet.column_dimensions["B"].width >= 10


def test_excel_bos_tabloda_gecerli_dosya_uretir(export_dir, bos_tablo):
    hedef = export_excel(bos_tablo, "bos.xlsx")
    sheet = load_workbook(hedef).active
    degerler = [cell.value for row in sheet.iter_rows() for cell in row]
    assert EMPTY_TABLE_MESSAGE in degerler
    assert hedef.stat().st_size > 0


def test_excel_zaman_dilimli_tarih_cokmez(export_dir):
    """Excel zaman dilimi tasimaz; aware datetime yazilmadan once soyulmali."""
    table = ReportTable(
        title="Zaman Damgasi",
        columns=[ReportColumn("an", "An", format="datetime")],
        rows=[{"an": datetime(2026, 8, 11, 7, 45, tzinfo=UTC)}],
    )
    hedef = export_excel(table, "zaman.xlsx")
    sheet = load_workbook(hedef).active
    baslik = _baslik_satiri(sheet, "An")
    hucre = sheet.cell(row=baslik + 1, column=1)
    assert isinstance(hucre.value, datetime)
    assert hucre.value.tzinfo is None


def test_excel_sayfa_adi_gecerli_karakterlere_indirgenir(export_dir):
    table = ReportTable(
        title="Gelir / Gider [2026] : Ozet - cok uzun bir rapor basligi",
        columns=[ReportColumn("a", "A")],
        rows=[{"a": "x"}],
    )
    hedef = export_excel(table, "ozet.xlsx")
    sheet = load_workbook(hedef).active
    assert len(sheet.title) <= 31
    assert not set(sheet.title) & set("/\\*?:[]")


# ==========================================================================
#  PDF
# ==========================================================================
def _pdf_sayfa_sayisi(data: bytes) -> int:
    """PDF icindeki sayfa nesnelerini sayar (``/Type /Pages`` haric)."""
    return len(re.findall(rb"/Type\s*/Page[^s]", data))


def test_pdf_olusur_ve_bos_degildir(export_dir, turkce_tablo):
    hedef = export_pdf(turkce_tablo, "gelir.pdf")
    ham = hedef.read_bytes()
    assert ham.startswith(b"%PDF")
    assert len(ham) > 1000


def test_pdf_turkce_karakterle_cokmez(export_dir, turkce_tablo):
    """Yerlesik Helvetica U+011F, U+015F, U+0131, U+0130 uretemez; TTF kaydi cozer."""
    hedef = export_pdf(turkce_tablo, "turkce.pdf")
    assert hedef.exists()
    assert hedef.stat().st_size > 0


def test_pdf_font_turkce_harfleri_kapsar():
    fonts = register_report_fonts()
    for harf in "ğĞşŞıİçÇöÖüÜ":
        assert fonts.supports(harf), f"Yazi tipi '{harf}' harfini kapsamiyor"


def test_pdf_desteklenmeyen_karakter_okunabilir_karsiliga_cevrilir():
    """Yazi tipi Turk Lirasi isaretini kapsamiyorsa 'TL' yazilir."""
    sinirli = ReportFonts("Helvetica", "Helvetica-Bold", frozenset(range(0x80)))
    assert sanitize_text("1.234,56 ₺", sinirli) == "1.234,56 TL"
    assert sanitize_text("", sinirli) == ""


def test_pdf_bilinmeyen_karakter_soru_isaretine_doner():
    sinirli = ReportFonts("Helvetica", "Helvetica-Bold", frozenset(range(0x80)))
    assert sanitize_text("A中B", sinirli) == "A?B"


def test_pdf_bos_tabloda_gecerli_dosya_uretir(export_dir, bos_tablo):
    hedef = export_pdf(bos_tablo, "bos.pdf")
    ham = hedef.read_bytes()
    assert ham.startswith(b"%PDF")
    assert _pdf_sayfa_sayisi(ham) == 1


def test_pdf_uzun_tablo_sayfalara_bolunur(export_dir):
    columns = [
        ReportColumn("sira", "Sira", align="right", format="integer"),
        ReportColumn("aciklama", "Aciklama"),
        ReportColumn("tutar", "Tutar", align="right", format="money"),
    ]
    rows = [
        {"sira": i, "aciklama": f"Örnek kayıt {i}", "tutar": Money.of(i * 10)}
        for i in range(1, 201)
    ]
    hedef = export_pdf(
        ReportTable(title="Uzun Rapor", columns=columns, rows=rows, footer_note=AI_NOTU),
        "uzun.pdf",
    )
    assert _pdf_sayfa_sayisi(hedef.read_bytes()) >= 2


def test_pdf_isaretleme_karakterleri_kacisla_yazilir(export_dir):
    """'<' ve '&' kacirilmazsa reportlab paragrafi ayristiramaz ve coker."""
    table = ReportTable(
        title="Kacis <testi> & ornek",
        columns=[ReportColumn("metin", "Metin")],
        rows=[{"metin": "A & B <script> tag"}],
    )
    hedef = export_pdf(table, "kacis.pdf")
    assert hedef.read_bytes().startswith(b"%PDF")


# ==========================================================================
#  Ortak ihracatci davranisi
# ==========================================================================
@pytest.mark.parametrize(
    ("ihracatci", "ad"),
    [(export_csv, "r.csv"), (export_excel, "r.xlsx"), (export_pdf, "r.pdf")],
)
def test_ihracatcilar_export_dizinine_yazar(export_dir, turkce_tablo, ihracatci, ad):
    hedef = ihracatci(turkce_tablo, ad)
    assert hedef.parent == export_dir.resolve()
    assert hedef.exists()


@pytest.mark.parametrize(
    ("ihracatci", "ad"),
    [(export_csv, "b.csv"), (export_excel, "b.xlsx"), (export_pdf, "b.pdf")],
)
def test_ihracatcilar_bos_tabloda_dosya_uretir(export_dir, bos_tablo, ihracatci, ad):
    hedef = ihracatci(bos_tablo, ad)
    assert hedef.exists()
    assert hedef.stat().st_size > 0


def test_alt_klasor_olusturulur(export_dir, turkce_tablo):
    hedef = export_csv(turkce_tablo, "2026/08/gelir.csv")
    assert hedef.exists()
    assert hedef.parent.name == "08"


def test_veri_kokunun_disina_yazilamaz(export_dir, turkce_tablo):
    with pytest.raises(ValidationError):
        export_csv(turkce_tablo, "../../../disari.csv")


def test_export_disindaki_klasore_yazilamaz(export_dir, turkce_tablo):
    """Veri koku icinde ama exports disinda kalan yollar da reddedilir."""
    with pytest.raises(ValidationError):
        export_csv(turkce_tablo, "../hotel.db")


def test_bilinmeyen_bicim_reddedilir():
    with pytest.raises(ValidationError):
        get_exporter("docx")


@pytest.mark.parametrize("bicim", ["csv", "CSV", ".xlsx", "excel", "pdf"])
def test_bicim_adindan_ihracatci_bulunur(bicim):
    assert callable(get_exporter(bicim))


# ==========================================================================
#  Sorgular - bos veritabani
# ==========================================================================
def test_bos_veritabaninda_tum_raporlar_bos_ama_gecerli(session: Session, sample_property):
    """Hicbir hareket yokken sorgular hata vermemeli, bos tablo dondurmeli."""
    tablolar = [
        queries.occupancy_report(session, sample_property.id, DONEM),
        queries.revenue_by_channel(session, sample_property.id, DONEM),
        queries.revenue_by_room_type(session, sample_property.id, DONEM),
        queries.revenue_by_charge_type(session, sample_property.id, DONEM),
        queries.daily_closing_report(session, sample_property.id, date(2026, 8, 11)),
        queries.arrivals_departures_report(session, sample_property.id, date(2026, 8, 11)),
        queries.housekeeping_report(session, sample_property.id, date(2026, 8, 11)),
        queries.maintenance_report(session, sample_property.id, DONEM),
        queries.stock_report(session, sample_property.id),
    ]
    for tablo in tablolar:
        assert tablo.column_count > 0
        assert tablo.title

    # Doluluk raporu gun basina bir satir uretir; digerleri bostur.
    assert len(tablolar[0].rows) == DONEM.nights
    assert all(tablo.is_empty for tablo in tablolar[1:])


def test_bos_veritabaninda_kpi_sifirdir(session: Session, sample_property):
    kpis = queries.kpi_report(session, sample_property.id, DONEM)
    assert kpis.room_nights_sold == 0
    assert kpis.occupancy_rate == 0.0
    assert kpis.adr.is_zero
    assert kpis.revpar.is_zero
    assert kpis.total_revenue.is_zero


def test_bos_rapor_ucu_de_disa_aktarilabilir(export_dir, session: Session, sample_property):
    tablo = queries.revenue_by_channel(session, sample_property.id, DONEM)
    tablo.footer_note = AI_NOTU
    assert export_csv(tablo, "bos-kanal.csv").exists()
    assert export_excel(tablo, "bos-kanal.xlsx").exists()
    assert export_pdf(tablo, "bos-kanal.pdf").exists()


def test_olmayan_folyo_ekstresi_hata_verir(session: Session):
    with pytest.raises(NotFoundError):
        queries.guest_ledger(session, 999_999)


# ==========================================================================
#  Sorgular - veri dolu
# ==========================================================================
def test_kpi_raporu_oda_gelirini_ayirir(session: Session, sample_property, dolu_veri):
    """ADR yalnizca ROOM ucretlerinden hesaplanmalidir."""
    kpis = queries.kpi_report(session, sample_property.id, DONEM)
    assert kpis.room_revenue == Money.of("3000.00")
    assert kpis.other_revenue == Money.of("800.00")  # restoran 500 + spa 300
    assert kpis.total_revenue == Money.of("3800.00")
    assert kpis.room_nights_sold == 3
    assert kpis.available_room_nights == 9  # 3 oda x 3 gece
    assert kpis.adr == Money.of("1000.00")
    assert kpis.revpar == Money.of("333.33")


def test_kpi_raporu_gecersiz_ucreti_saymaz(session: Session, sample_property, dolu_veri):
    """is_void ucret satiri gelire girmemeli (denetim izi icin silinmez)."""
    kpis = queries.kpi_report(session, sample_property.id, DONEM)
    assert kpis.room_revenue == Money.of("3000.00")


def test_arizali_oda_kpi_paydasini_dusurur(
    session: Session, sample_property, sample_rooms, dolu_veri
):
    sample_rooms[2].housekeeping_status = RoomHousekeepingStatus.OUT_OF_ORDER
    session.commit()
    kpis = queries.kpi_report(session, sample_property.id, DONEM)
    assert kpis.available_room_nights == 6
    assert kpis.revpar == Money.of("500.00")


def test_doluluk_raporu_gun_bazinda_satir_uretir(session: Session, sample_property, dolu_veri):
    tablo = queries.occupancy_report(session, sample_property.id, DONEM)
    assert len(tablo.rows) == 3
    assert tablo.rows[0]["dolu_oda"] == 1
    assert tablo.rows[0]["satilabilir_oda"] == 3
    assert tablo.rows[0]["doluluk"] == pytest.approx(33.33, abs=0.01)


def test_ucret_turu_raporu_gruplar(session: Session, sample_property, dolu_veri):
    tablo = queries.revenue_by_charge_type(session, sample_property.id, DONEM)
    turler = {row["ucret_turu"] for row in tablo.rows}
    assert ChargeType.ROOM in turler
    assert ChargeType.RESTAURANT in turler
    assert ChargeType.SPA in turler
    oda = next(row for row in tablo.rows if row["ucret_turu"] is ChargeType.ROOM)
    assert oda["toplam"] == Money.of("3000.00")
    assert oda["adet"] == 3


def test_kanal_raporu_rezervasyon_kaynagina_gore_gruplar(
    session: Session, sample_property, dolu_veri
):
    tablo = queries.revenue_by_channel(session, sample_property.id, DONEM)
    assert len(tablo.rows) == 1
    assert tablo.rows[0]["kanal"] is ReservationSource.BOOKING_COM
    assert tablo.rows[0]["toplam"] == Money.of("3800.00")


def test_oda_tipi_raporu_folyo_baglantisi_olmayani_da_toplar(
    session: Session, sample_property, dolu_veri
):
    """Oda satirina bagli olmayan folyo, rezervasyonun ilk odasina yazilir."""
    tablo = queries.revenue_by_room_type(session, sample_property.id, DONEM)
    assert len(tablo.rows) == 1
    satir = tablo.rows[0]
    assert satir["oda_tipi"] == "Standart Oda"
    assert satir["oda_geliri"] == Money.of("3000.00")
    assert satir["diger_gelir"] == Money.of("800.00")


def test_gun_sonu_raporu_gelir_tahsilat_ve_kasa_icerir(
    session: Session, sample_property, dolu_veri
):
    tablo = queries.daily_closing_report(session, sample_property.id, date(2026, 8, 12))
    gruplar = {row["grup"] for row in tablo.rows}
    assert {"Gelir", "Tahsilat", "Kasa", "Ozet"} <= gruplar
    ozet = {row["kalem"]: row["tutar"] for row in tablo.rows if row["grup"] == "Ozet"}
    assert ozet["Net Tahsilat"] == Money.of("2000.00")


def test_giris_cikis_raporu_turleri_ayirir(session: Session, sample_property, dolu_veri):
    giris = queries.arrivals_departures_report(session, sample_property.id, date(2026, 8, 10))
    ara_gun = queries.arrivals_departures_report(session, sample_property.id, date(2026, 8, 11))
    cikis = queries.arrivals_departures_report(session, sample_property.id, date(2026, 8, 13))

    assert [row["tur"] for row in giris.rows] == ["Giris"]
    assert [row["tur"] for row in ara_gun.rows] == ["Konaklama"]
    assert [row["tur"] for row in cikis.rows] == ["Cikis"]
    assert giris.rows[0]["oda"] == "101"


def test_misafir_ekstresi_yuruyen_bakiye_hesaplar(session: Session, dolu_veri):
    tablo = queries.guest_ledger(session, dolu_veri["folio"].id)
    # 3 oda ucreti + 1 restoran + 1 tahsilat = 5 satir (gecersiz kilinan haric)
    assert len(tablo.rows) == 5
    assert tablo.rows[-1]["bakiye"] == Money.of("1500.00")  # 3500 borc - 2000 tahsilat


def test_kat_hizmetleri_raporu(session: Session, sample_property, dolu_veri):
    tablo = queries.housekeeping_report(session, sample_property.id, date(2026, 8, 11))
    assert len(tablo.rows) == 1
    assert tablo.rows[0]["oda"] == "101"
    assert tablo.rows[0]["personel"] == "Atanmadi"


def test_teknik_servis_raporu(session: Session, sample_property, dolu_veri):
    tablo = queries.maintenance_report(session, sample_property.id, DONEM)
    assert len(tablo.rows) == 1
    assert tablo.rows[0]["fis_no"] == "ARZ-0001"
    assert tablo.rows[0]["maliyet"] == Money.of("150.00")


def test_stok_raporu_kritik_satirlari_one_alir(session: Session, sample_property, dolu_veri):
    tablo = queries.stock_report(session, sample_property.id)
    assert len(tablo.rows) == 1
    assert tablo.rows[0]["durum"] == "Kritik"
    assert tablo.rows[0]["stok_degeri"] == Money.of("50.00")


def test_dolu_rapor_ucu_de_disa_aktarilabilir(
    export_dir, session: Session, sample_property, dolu_veri
):
    tablo = queries.revenue_by_charge_type(session, sample_property.id, DONEM)
    tablo.footer_note = AI_NOTU
    csv_yolu = export_csv(tablo, "ucret.csv")
    assert "Oda Ucreti" in csv_yolu.read_text(encoding="utf-8-sig")
    assert export_excel(tablo, "ucret.xlsx").stat().st_size > 0
    assert export_pdf(tablo, "ucret.pdf").read_bytes().startswith(b"%PDF")


def test_kpi_tablosu_disa_aktarilabilir(export_dir, session: Session, sample_property, dolu_veri):
    tablo = queries.kpi_report(session, sample_property.id, DONEM).to_table()
    hedef = export_csv(tablo, "kpi.csv")
    metin = hedef.read_text(encoding="utf-8-sig")
    assert "RevPAR" in metin
    assert "ADR" in metin


# ==========================================================================
#  Gerileme testleri - gecmis donem doluluk
# ==========================================================================
def _konaklama_ekle(
    session: Session,
    sample_property,
    sample_rooms,
    sample_guest,
    *,
    durum: ReservationStatus,
    onay_no: str,
    oda_sirasi: int = 0,
) -> None:
    """Uc gecelik, tek odali, oda ucreti islenmis bir konaklama ekler."""
    reservation = Reservation(
        property_id=sample_property.id,
        confirmation_number=onay_no,
        status=durum,
        source=ReservationSource.DIRECT,
        primary_guest_id=sample_guest.id,
        check_in_date=date(2026, 8, 10),
        check_out_date=date(2026, 8, 13),
        adults=2,
        currency=Currency.TRY,
    )
    session.add(reservation)
    session.flush()

    res_room = ReservationRoom(
        reservation_id=reservation.id,
        room_type_id=sample_rooms[oda_sirasi].room_type_id,
        room_id=sample_rooms[oda_sirasi].id,
        check_in_date=date(2026, 8, 10),
        check_out_date=date(2026, 8, 13),
        adults=2,
        nightly_rate=Decimal("1000.00"),
        total_amount=Decimal("3000.00"),
    )
    session.add(res_room)
    session.flush()

    folio = Folio(
        property_id=sample_property.id,
        folio_number=f"F-{onay_no}",
        reservation_id=reservation.id,
        reservation_room_id=res_room.id,
        guest_id=sample_guest.id,
    )
    session.add(folio)
    session.flush()

    for gun in (date(2026, 8, 10), date(2026, 8, 11), date(2026, 8, 12)):
        charge = Charge(
            folio_id=folio.id,
            charge_type=ChargeType.ROOM,
            description="Konaklama",
            charge_date=gun,
            quantity=Decimal("1.000"),
            unit_price=Decimal("1000.00"),
            tax_rate_percent=Decimal("0.00"),
        )
        charge.compute_totals()
        session.add(charge)
    session.commit()


def test_cikis_yapmis_konaklama_gecmis_donem_raporunda_sayilir(
    session: Session, sample_property, sample_rooms, sample_guest
):
    """Gecen ayin raporu CHECKED_OUT kayitlardan olusur - gorunmez olamaz.

    Yalnizca "envanteri bloke eden" durumlara bakan bir sorgu, her gecmis
    donem icin doluluk %0 ve ADR 0 uretir; gelir satirlari doluyken.
    """
    _konaklama_ekle(
        session,
        sample_property,
        sample_rooms,
        sample_guest,
        durum=ReservationStatus.CHECKED_OUT,
        onay_no="GECMIS-01",
    )

    kpis = queries.kpi_report(session, sample_property.id, DONEM)
    assert kpis.room_nights_sold == 3
    assert kpis.available_room_nights == 9
    assert kpis.room_revenue == Money.of("3000.00")
    assert kpis.adr == Money.of("1000.00")
    assert kpis.occupancy_rate > 0

    tablo = queries.occupancy_report(session, sample_property.id, DONEM)
    assert [row["dolu_oda"] for row in tablo.rows] == [1, 1, 1]


def test_iptal_ve_gelmeme_dolulugu_sisirmez(
    session: Session, sample_property, sample_rooms, sample_guest
):
    """Iptal veya no-show durumunda oda satilmamistir; doluluga girmemeli."""
    _konaklama_ekle(
        session,
        sample_property,
        sample_rooms,
        sample_guest,
        durum=ReservationStatus.CANCELLED,
        onay_no="IPTAL-01",
        oda_sirasi=0,
    )
    _konaklama_ekle(
        session,
        sample_property,
        sample_rooms,
        sample_guest,
        durum=ReservationStatus.NO_SHOW,
        onay_no="GELMEDI-01",
        oda_sirasi=1,
    )

    kpis = queries.kpi_report(session, sample_property.id, DONEM)
    assert kpis.room_nights_sold == 0
    assert kpis.occupancy_rate == 0.0
    # Iptal ve gelmeme oranlarinin paydasi yine de iki rezervasyondur.
    assert kpis.cancellation_rate == 0.5
    assert kpis.no_show_rate == 0.5

    tablo = queries.occupancy_report(session, sample_property.id, DONEM)
    assert [row["dolu_oda"] for row in tablo.rows] == [0, 0, 0]


def test_oda_tipi_raporu_rezervasyonsuz_folyoyu_kaybetmez(
    session: Session, sample_property, sample_guest, dolu_veri
):
    """Rezervasyona bagli olmayan folyo ucretleri rapordan dusmemeli.

    Ic birlestirme kullanildiginda bu satirlar sessizce kayboluyor ve ayni
    donemin oda tipi raporu ile ucret turu raporu farkli toplam veriyordu.
    """
    bar_folyo = Folio(
        property_id=sample_property.id,
        folio_number="F-BAR-01",
        guest_id=sample_guest.id,
    )
    session.add(bar_folyo)
    session.flush()
    bar_ucreti = Charge(
        folio_id=bar_folyo.id,
        charge_type=ChargeType.FOOD_BEVERAGE,
        description="Bar hesabi",
        charge_date=date(2026, 8, 11),
        quantity=Decimal("1.000"),
        unit_price=Decimal("450.00"),
        tax_rate_percent=Decimal("0.00"),
    )
    bar_ucreti.compute_totals()
    session.add(bar_ucreti)
    session.commit()

    oda_tipi_tablosu = queries.revenue_by_room_type(session, sample_property.id, DONEM)
    ucret_turu_tablosu = queries.revenue_by_charge_type(session, sample_property.id, DONEM)

    tipler = {row["oda_tipi"] for row in oda_tipi_tablosu.rows}
    assert "Belirtilmemis" in tipler

    oda_tipi_toplami = sum(
        (row["toplam"] for row in oda_tipi_tablosu.rows), Money.zero(Currency.TRY)
    )
    ucret_turu_toplami = sum(
        (row["toplam"] for row in ucret_turu_tablosu.rows), Money.zero(Currency.TRY)
    )
    assert oda_tipi_toplami == ucret_turu_toplami == Money.of("4250.00")


# ==========================================================================
#  Formul enjeksiyonu
# ==========================================================================
@pytest.fixture
def zararli_tablo() -> ReportTable:
    """Misafir adi alanina formul yazilmis (uydurma) bir kayit."""
    return ReportTable(
        title="Misafir Listesi",
        columns=[
            ReportColumn("misafir", "Misafir"),
            ReportColumn("tutar", "Tutar", align="right", format="money"),
        ],
        rows=[
            {"misafir": "=1+1", "tutar": Money.of("100.00")},
            {"misafir": "@SUM(A1:A9)", "tutar": Money.of("-250.00")},
            {"misafir": "Normal Misafir", "tutar": Money.of("50.00")},
        ],
    )


def test_csv_formul_enjeksiyonu_etkisizlestirilir(export_dir, zararli_tablo):
    """Misafir adina yazilan formul, Excel'de calisan bir hucreye donusmemeli."""
    hedef = export_csv(zararli_tablo, "zararli.csv")
    with hedef.open(encoding="utf-8-sig", newline="") as handle:
        satirlar = list(csv.reader(handle, delimiter=CSV_DELIMITER))
    assert satirlar[1][0] == "'=1+1"
    assert satirlar[2][0] == "'@SUM(A1:A9)"
    assert satirlar[3][0] == "Normal Misafir"


def test_csv_negatif_tutar_bozulmaz(export_dir, zararli_tablo):
    """Eksi bakiye formul sayilmaz; tirnak eklemek tutari metne cevirirdi."""
    hedef = export_csv(zararli_tablo, "zararli.csv")
    with hedef.open(encoding="utf-8-sig", newline="") as handle:
        satirlar = list(csv.reader(handle, delimiter=CSV_DELIMITER))
    assert satirlar[2][1].startswith("-250,00")


def test_excel_formul_enjeksiyonu_metin_olarak_yazilir(export_dir, zararli_tablo):
    """openpyxl '=' ile baslayan metni formul isaretler; bu ezilmelidir."""
    hedef = export_excel(zararli_tablo, "zararli.xlsx")
    sheet = load_workbook(hedef).active
    baslik = _baslik_satiri(sheet, "Misafir")
    for offset in (1, 2):
        hucre = sheet.cell(row=baslik + offset, column=1)
        assert hucre.data_type == "s", f"{hucre.value!r} formul olarak yazilmis"


def test_excel_ust_bilgi_de_formul_olmaz(export_dir):
    """Alt baslik misafir adi tasiyabilir (misafir ekstresi); o da metindir."""
    table = ReportTable(
        title="Misafir Hesap Ekstresi",
        columns=[ReportColumn("a", "A")],
        rows=[{"a": "x"}],
        subtitle='=HYPERLINK("http://ornek-test.local")',
    )
    hedef = export_excel(table, "ustbilgi.xlsx")
    sheet = load_workbook(hedef).active
    formuller = [
        cell.value
        for row in sheet.iter_rows()
        for cell in row
        if getattr(cell, "data_type", None) == "f"
    ]
    assert formuller == []
