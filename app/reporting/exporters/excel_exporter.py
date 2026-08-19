"""Excel (``.xlsx``) ihracatcisi - openpyxl tabanli.

Neden ham deger yaziliyor?
--------------------------
CSV ve PDF ciktilarinda hucreler metne cevrilir. Excel'de ise **sayilar
sayi olarak** yazilir ve gorunum ``number_format`` ile ayarlanir. Sebep
basit: kullanici raporu acip toplam almak, suzmek, grafik cizmek ister.
``"1.234,56 ₺"`` metnini iceren bir sutunda bunlarin hicbiri calismaz.

Zaman dilimi tuzagi
-------------------
Excel'in tarih-saat bicimi zaman dilimi tasimaz; openpyxl, ``tzinfo``
tasiyan bir ``datetime`` yazilmak istendiginde
``ValueError: Excel does not support timezones in datetimes`` firlatir.
Uygulama genelinde tum zaman damgalari UTC-aware oldugu icin (bkz.
:class:`app.infrastructure.db.types.TZDateTime`) bu hata *her* rapor
ihracatinda ortaya cikardi. :func:`_excel_value` degeri yazmadan once
``tzinfo``'yu duserek sorunu kaynaginda cozer.

Formul enjeksiyonu
------------------
openpyxl, ``=`` ile baslayan bir metni **formul** olarak isaretler
(``data_type == "f"``). Misafir adi ya da ucret aciklamasi gibi kullanici
girdisi rapora dogrudan girdigi icin, bu davranis dosyayi acan personelin
makinesinde kod calistirilmasina kapi acardi. :func:`_write_text` her metin
hucresini acikca metin (``"s"``) olarak isaretler; boylece icerik ne olursa
olsun Excel onu veri sayar. CSV tarafinda ayni sorun tek tirnak on ekiyle
cozulur - orada hucre tipi diye bir kavram yoktur.
"""

from __future__ import annotations

import os
import re
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.core.log import get_logger
from app.domain.enums import LabeledEnum
from app.domain.value_objects import Money
from app.reporting.models import (
    DATETIME_FORMAT,
    ReportColumn,
    ReportTable,
    format_cell,
    resolve_export_path,
)

log = get_logger(__name__)

#: Baslik satirinin dolgu rengi (koyu mavi) ve yazi rengi.
HEADER_FILL = "1F4E79"
HEADER_FONT_COLOR = "FFFFFF"

#: Excel sutun genisligi sinirlari - cok dar sutun okunmaz, cok genis sutun
#: yazdirmayi bozar.
MIN_COLUMN_WIDTH = 10
MAX_COLUMN_WIDTH = 55

#: Excel sayfa adinda kullanilamayan karakterler.
_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")

#: Sutun turune gore Excel bicim kodu.
_NUMBER_FORMATS: dict[str, str] = {
    "integer": "#,##0",
    "decimal": "#,##0.00",
    "money": "#,##0.00",
    "percent": '0.00"%"',
    "date": "DD.MM.YYYY",
    "datetime": "DD.MM.YYYY HH:MM",
}


def _sheet_title(title: str) -> str:
    """Rapor basligini gecerli bir Excel sayfa adina cevirir.

    Excel sayfa adlari en fazla 31 karakterdir ve ``\\ / * ? : [ ]``
    karakterlerini kabul etmez; aksi halde dosya olusturulur ama Excel
    "onarilmasi gerekiyor" uyarisi verir.
    """
    cleaned = _INVALID_SHEET_CHARS.sub("-", title).strip() or "Rapor"
    return cleaned[:31]


def _excel_value(value: Any, column: ReportColumn) -> Any:
    """Bir degeri Excel'in anlayacagi ilkel tipe cevirir."""
    if value is None:
        return None
    if isinstance(value, Money):
        # Decimal openpyxl tarafindan yerel olarak desteklenir; float'a
        # cevirmek kurus kaybina yol acabilirdi.
        return value.amount
    if isinstance(value, LabeledEnum):
        return value.label
    if isinstance(value, bool):
        return "Evet" if value else "Hayir"
    if isinstance(value, datetime):
        # Excel zaman dilimi tasimaz; aware deger oldugu gibi yazilamaz.
        return value.replace(tzinfo=None)
    if isinstance(value, (int, float, Decimal, date)):
        return value
    return format_cell(value, column)


def _mark_as_text(cell: Cell) -> None:
    """Metin hucresini formul olarak yorumlanmaktan korur.

    openpyxl deger atandiginda tipi kendisi tahmin eder ve ``=`` ile
    baslayan her metni formul sayar. Tahmini yazma sonrasi acikca ezmek,
    degeri kirpmadan (yani rapor icerigini bozmadan) enjeksiyonu keser.
    """
    if cell.data_type == "f":
        cell.data_type = "s"


def _number_format(value: Any, column: ReportColumn) -> str | None:
    """Hucreye uygulanacak Excel bicim kodu."""
    if isinstance(value, Money):
        return f'#,##0.00 "{value.currency.symbol}"'
    if isinstance(value, datetime):
        return _NUMBER_FORMATS["datetime"]
    if isinstance(value, date):
        return _NUMBER_FORMATS["date"]
    return _NUMBER_FORMATS.get(column.format)


def _auto_width(sheet: Worksheet, table: ReportTable) -> None:
    """Sutun genisliklerini icerige gore ayarlar.

    Genislik, **bicimlenmis metin** uzunluguna gore hesaplanir; ham
    ``Decimal`` uzunlugu ("1234.56") gercekte gorunecek metinden
    ("1.234,56 ₺") kisadir ve sutunlar dar kalirdi.
    """
    for index, column in enumerate(table.columns, start=1):
        widest = len(column.title)
        for row in table.rows:
            widest = max(widest, len(format_cell(row.get(column.key), column)))
        width = min(max(widest + 3, MIN_COLUMN_WIDTH), MAX_COLUMN_WIDTH)
        sheet.column_dimensions[get_column_letter(index)].width = width


def export(table: ReportTable, output_path: str | os.PathLike[str]) -> Path:
    """Rapor tablosunu ``.xlsx`` olarak yazar ve dosya yolunu dondurur.

    Cikti: ust bilgi bloku (baslik, alt baslik, suzgecler, uretim zamani),
    kalin ve renkli sutun basliklari, dondurulmus baslik satiri, otomatik
    suzgec ve icerige gore sutun genislikleri.
    """
    target = resolve_export_path(output_path)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _sheet_title(table.title)

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---- Ust bilgi bloku ----
    # Alt baslik da kullanici verisi tasiyabilir (or. misafir ekstresinde
    # misafir adi), bu yuzden ust bilgi hucreleri de metne sabitlenir.
    row_index = 1
    title_cell = sheet.cell(row=row_index, column=1, value=table.title)
    title_cell.font = Font(bold=True, size=14)
    _mark_as_text(title_cell)
    row_index += 1
    for text in (
        table.subtitle,
        table.filters_description,
        f"Olusturma: {table.generated_at.strftime(DATETIME_FORMAT)} (UTC)",
    ):
        if not text:
            continue
        info_cell = sheet.cell(row=row_index, column=1, value=text)
        info_cell.font = Font(italic=True, size=10)
        _mark_as_text(info_cell)
        row_index += 1
    row_index += 1  # bos ayirici satir

    header_row = row_index

    # ---- Baslik satiri ----
    fill = PatternFill("solid", fgColor=HEADER_FILL)
    header_font = Font(bold=True, color=HEADER_FONT_COLOR)
    for index, column in enumerate(table.columns, start=1):
        cell = sheet.cell(row=header_row, column=index, value=column.title)
        cell.fill = fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _mark_as_text(cell)

    # ---- Veri satirlari ----
    if table.is_empty:
        for index, text in enumerate(table.empty_row(), start=1):
            cell = sheet.cell(row=header_row + 1, column=index, value=text or None)
            cell.border = border
            cell.font = Font(italic=True)
            _mark_as_text(cell)
        last_row = header_row + 1
    else:
        for offset, row in enumerate(table.rows, start=1):
            for index, column in enumerate(table.columns, start=1):
                raw = row.get(column.key)
                cell = sheet.cell(row=header_row + offset, column=index)
                cell.value = _excel_value(raw, column)
                number_format = _number_format(raw, column)
                if number_format:
                    cell.number_format = number_format
                cell.alignment = Alignment(horizontal=column.align, vertical="top")
                cell.border = border
                _mark_as_text(cell)
        last_row = header_row + len(table.rows)

    # ---- Dondurulmus baslik + oto filtre ----
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)
    if table.column_count:
        last_column = get_column_letter(table.column_count)
        sheet.auto_filter.ref = f"A{header_row}:{last_column}{last_row}"

    _auto_width(sheet, table)

    if table.footer_note:
        note = sheet.cell(row=last_row + 2, column=1, value=table.footer_note)
        note.font = Font(italic=True, size=9, color="808080")
        _mark_as_text(note)

    workbook.save(target)
    log.info("rapor_disa_aktarildi", bicim="xlsx", satir=len(table.rows))
    return target


__all__ = [
    "HEADER_FILL",
    "MAX_COLUMN_WIDTH",
    "MIN_COLUMN_WIDTH",
    "export",
]
